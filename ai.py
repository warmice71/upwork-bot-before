from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import *
from time import sleep
from datetime import datetime
import sys
import os
import json
import sqlite3
import random
import openpyxl


def waitInfinite(callback, debug=False, callNum=20):
    sleep(1)
    for i in range(callNum):
        try:
            callback()
            break
        except NoSuchElementException as e:
            print("{} on line {}".format(str(e).split(
                '\n')[0], sys.exc_info()[-1].tb_lineno))
            sleep(1)
            pass
        except JavascriptException as e:
            print("{} on line {}".format(str(e).split(
                '\n')[0], sys.exc_info()[-1].tb_lineno))
            sleep(1)
            pass
        except StaleElementReferenceException as e:
            print("{} on line {}".format(str(e).split(
                '\n')[0], sys.exc_info()[-1].tb_lineno))
            sleep(1)
            pass
        except ElementClickInterceptedException as e:
            print("{} on line {}".format(str(e).split(
                '\n')[0], sys.exc_info()[-1].tb_lineno))
            sleep(1)
            pass
        except ElementNotInteractableException as e:
            print("{} on line {}".format(str(e).split(
                '\n')[0], sys.exc_info()[-1].tb_lineno))
            sleep(1)
            pass
        except Exception as e:
            print("{} on line {}".format(str(e).split(
                '\n')[0], sys.exc_info()[-1].tb_lineno))
            # sleep(1)
            # pass
            # driver.quit()
            # exit()


def waitUntil(callback, driver, selector):
    sleep(1)
    yet = True
    while yet:
        try:
            callback(driver.execute_script(
                "x=document.querySelectorAll('{}').length;return document.querySelectorAll('{}')[x-1]".format(selector, selector)))
            yet = False
        except Exception as e:
            print(str(e).split('\n')[0])
            sleep(1)
            pass


def waitUntil1(callback, driver, selector):
    sleep(1)
    yet = True
    while yet:
        try:
            callback(driver.find_elements(By.CSS_SELECTOR, selector)[3])
            yet = False
        except Exception as e:
            print(str(e).split('\n')[0])
            sleep(1)
            pass


def clickByMouse(element):
    ActionChains(driver).click(element)\
                        .perform()


def selectDropDown(itemSelector, country):
    nations = driver.find_elements(By.CSS_SELECTOR, itemSelector)

    if str(type(country)) == "<class 'int'>":
        driver.execute_script(
            f'document.querySelectorAll("{itemSelector}")[{str(country)}].click()')
    else:
        for i in range(len(nations)):
            try:
                if nations[i].text.find(country) >= 0:
                    driver.execute_script(
                        f'document.querySelectorAll("{itemSelector}")[{str(i)}].click()')
                    break
            except:
                pass


def selectDateDropDown(dropdownId, itemSelector, country):
    tmp = dropdownId.split('##')
    if len(tmp) == 2:
        dropdownId = tmp[0]
        driver.execute_script(
            f'document.querySelectorAll(\'div[aria-labelledby^="{dropdownId}"]\')[{tmp[1]}].click()')
    else:
        driver.execute_script(
            f'document.querySelector(\'div[aria-labelledby^="{dropdownId}"]\').click()')
    sleep(1)
    nations = driver.find_elements(By.CSS_SELECTOR, itemSelector)

    if str(type(country)) == "<class 'int'>":
        driver.execute_script(
            f'document.querySelectorAll("{itemSelector}")[{str(country)}].click()')
    else:
        for i in range(len(nations)):
            try:
                if nations[i].text.find(country) >= 0:
                    driver.execute_script(
                        f'document.querySelectorAll("{itemSelector}")[{str(i)}].click()')
                    break
            except:
                pass


def addExperience(driver, experience):
    global expFlag
    if expFlag:
        waitInfinite(lambda: driver.execute_script(
            'document.querySelectorAll("button.air3-btn.air3-btn-circle")[0].click()'))
    else:
        waitInfinite(lambda: driver.execute_script(
            'document.querySelector("button.air3-btn.air3-btn-secondary.air3-btn-circle").click()'))

    sleep(3)

    if expFlag:
        waitInfinite(lambda: driver.find_elements(By.CSS_SELECTOR,
                     'input[aria-labelledby="title-label"]')[1].send_keys(""))
    waitInfinite(lambda: driver.find_elements(By.CSS_SELECTOR,
                 'input[aria-labelledby="title-label"]')[1].send_keys(experience['role']))
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[aria-labelledby="company-label"]').send_keys(experience['company']))
    # waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR, 'input[aria-labelledby="location-label"]').send_keys(city))
    # waitInfinite(lambda: selectDateDropDown("location-label", "span.air3-menu-item-text", country))

    sleep(1)

    start_year = eval(experience['start'].split('.')[0])
    start_month = eval(experience['start'].split('.')[1])
    waitInfinite(lambda: selectDateDropDown("start-date-month",
                 "span.air3-menu-item-text", start_month - 1))
    waitInfinite(lambda: selectDateDropDown("start-date-year",
                 "span.air3-menu-item-text", curr_year - start_year))

    # waitInfinite(lambda: driver.find_element(By.TAG_NAME,'textarea').send_keys(description))
    # sleep(3)
    for text in experience['description']:
        waitInfinite(lambda: driver.find_element(
            By.CSS_SELECTOR, 'textarea[aria-labelledby="description-label"]').send_keys("• " + text + "\n"))

    if experience['end'] == 'current':
        driver.execute_script(
            'document.querySelector(\'[data-qa="currently-working"]\').querySelector("label").click()')
    else:
        end_year = eval(experience['end'].split('.')[0])
        end_month = eval(experience['end'].split('.')[1])
        waitInfinite(lambda: selectDateDropDown("end-date-month",
                     "span.air3-menu-item-text", end_month - 1))
        waitInfinite(lambda: selectDateDropDown("end-date-year",
                     "span.air3-menu-item-text", curr_year - end_year))

    # debug()

    driver.execute_script(
        'document.querySelector(\'button[data-qa="btn-save"]\').click()')

    expFlag = False


def addEducation(driver, education):
    start = eval(education['start'])
    end = eval(education['end'])
    global expFlag
    # if expFlag:
    #     driver.find_element(By.XPATH, "/html/body/div[5]/div/div[1]/div/div[0]/div[1]/div/div[3]/div/div[1]/div[0]/a/div/button").click()
    #     # driver.execute_script('document.querySelector(\'button[data-qa="education-add-btn"]\').click()')
    # else:
    # waitInfinite(lambda: driver.execute_script('document.querySelectorAll("button.air3-btn.air3-btn-secondary.air3-btn-circle")[1].click()'))
    waitInfinite(lambda: driver.execute_script(
        'document.querySelector(\'button[data-qa="education-add-btn"]\').click()'))

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[aria-labelledby="school-label"]').click())
    sleep(1)
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[aria-labelledby="school-label"]').send_keys(education['university']))
    sleep(1)
    try:
        driver.find_element(By.CLASS_NAME, "air3-menu-item-text").click()
    except:
        pass

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[aria-labelledby="degree-label"]').click())
    sleep(1)
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[aria-labelledby="degree-label"]').send_keys(education['degree']))
    sleep(1)

    action.move_to_element_with_offset(driver.find_element(
        By.CSS_SELECTOR, 'input[aria-labelledby="degree-label"]'), 50, 70)
    action.click()
    action.perform()
    # debug()
    # driver.find_element(By.CLASS_NAME, "is-focused").click()
    # driver.find_element(By.CLASS_NAME, "air3-menu-item-text").click()

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[aria-labelledby="area-of-study-label"]').click())
    sleep(1)
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[aria-labelledby="area-of-study-label"]').send_keys(education['field']))
    sleep(1)
    action.move_to_element_with_offset(driver.find_element(
        By.CSS_SELECTOR, 'input[aria-labelledby="area-of-study-label"]'), 50, 70)
    action.click()
    action.perform()

    waitInfinite(lambda: selectDateDropDown("dates-attended-label##0",
                 "span.air3-menu-item-text", curr_year - start + 1+5))
    waitInfinite(lambda: selectDateDropDown(
        "dates-attended-label##1", "span.air3-menu-item-text", 2030 - end + 1+5))
    sleep(3)
    driver.execute_script(
        'document.querySelector(\'button[data-qa="btn-save"]\').click()')

    expFlag = False


def addLanguage(driver, language, pro, count):
    pro = eval(pro)
    driver.execute_script(
        'document.querySelector("button.air3-btn.air3-btn-secondary.air3-btn-sm").click()')
    sleep(2)

    waitInfinite(lambda: selectDateDropDown(
        f"dropdown-label-language-{count}", "span.air3-menu-item-text", language))
    waitInfinite(lambda: selectDateDropDown(
        f"dropdown-label-proficiency-{count}", "span.air3-menu-item-text", pro))

    return count + 1


def addSkill(driver, inp, skill, field="skills-input", end=2):

    waitUntil(lambda x: x.click(), driver, f'input[aria-labelledby="{field}"]')
    for i in skill:
        driver.find_element(
            By.CSS_SELECTOR, f'input[aria-labelledby="{field}"]').send_keys(i)
        sleep(0.3)

    flag = True
    while flag:
        nations = driver.find_elements(
            By.CSS_SELECTOR, "span.air3-menu-item-text")
        flag = len(nations) == 0

    for i in range(len(nations)):
        try:
            if nations[i].text.find(skill) >= 0:
                driver.execute_script(
                    f'document.querySelectorAll("span.air3-menu-item-text")[{str(i)}].click()')
                break
        except:
            pass
    waitUntil(lambda x: x.clear(), driver, f'input[aria-labelledby="{field}"]')


def addService(driver, services):
    waitUntil(lambda x: x.click(), driver, 'div[data-test="dropdown-toggle"]')
    sleep(1)

    for service in services:
        driver.execute_script(f'''
            // document.querySelectorAll(\'div[data-test="dropdown-toggle"]\')[3].click()
            var services = document.querySelectorAll('span.air3-menu-checkbox-labels');
            var toselect;
            for (let i = 0; i < services.length; i++) {{
                console.log(services[i], '{service}');
                if (services[i].textContent.indexOf('{service}') >= 0) {{
                    toselect = services[i].parentNode.parentNode;
                    break;
                }}
            }}
            if (toselect) {{
                if (toselect.getAttribute("aria-selected") == 'false') {{
                    toselect.parentNode.parentNode.parentNode.click();
                    setTimeout(() => toselect.click(), 300);
                }}
            }}
        ''')
        sleep(0.1)


def configLast(driver, country, street, city, birthday, phone, photo):
    # selectDateDropDown("country-label", "span.air3-menu-item-text", country)
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[aria-labelledby="street-label"]').send_keys(street))

    addSkill(driver, driver.find_element(By.CSS_SELECTOR,
             'input[aria-labelledby="city-label"]'), city, "city-label", end=3)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[aria-labelledby="date-of-birth-label"]').send_keys(birthday))
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[aria-labelledby^="dropdown-label-phone-number"]').send_keys(phone))
    # waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR, 'input[aria-labelledby^="state-label"]').send_keys(state))
    sleep(1)
    waitInfinite(lambda: driver.execute_script(
        "document.querySelector('button[data-qa=\"open-loader\"]').click()"))
    sleep(1)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR, 'input[type="file"]').send_keys(
        os.path.join(os.getcwd(), 'users', photo)))
    waitInfinite(lambda: driver.execute_script(
        "document.querySelectorAll('button.air3-btn.air3-btn-primary')[document.querySelectorAll('button.air3-btn.air3-btn-primary').length - 1].click()"))

    # selectDateDropDown("country-label", "span.air3-menu-item-text", country)
    # waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR, 'input[aria-labelledby="street-label"]').send_keys(street))

    # addSkill(driver, driver.find_element(By.CSS_SELECTOR, 'input[aria-labelledby="city-label"]'), city, "city-label")

    # waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR, 'input[aria-labelledby="postal-code-label"]').send_keys(zipcode))
    # waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR, 'input[aria-labelledby^="dropdown-label-phone-number"]').send_keys(phone))

    # waitInfinite(lambda: driver.execute_script("document.querySelector('button[data-cy=\"open-loader\"]').click()"))
    # sleep(1)

    # waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR, 'input[type="file"]').send_keys(os.path.join(os.getcwd(), 'users', photo)))
    # waitInfinite(lambda: driver.execute_script("document.querySelectorAll('button.air3-btn.air3-btn-primary')[5].click()"))


print("#"*100)
global expFlag, tempURL
expFlag = True

curr_year = datetime.now().year
# name = sys.argv[1]

# user settings
name = 'ai'
pwd = "sjh030701"


with open('./users/' + name + '.json', 'r', encoding="utf-8-sig") as f:
    profile = json.load(f)

emailFile = './emails_ai.xlsx'
excel = openpyxl.load_workbook(emailFile)
sheet = excel.active
rowCount = sheet.max_row

finish_excel = openpyxl.Workbook()
sht = finish_excel.active
file_name = "./finished.xlsx"

for p in range(1, rowCount+1):

    email = sheet.cell(row=p, column=1).value
    print(email)

    sleep(1)

    options = webdriver.ChromeOptions()

    options.add_argument('--start-maximized')
    # options.add_extension(os.path.join(os.getcwd(), 'extensions', 'adblock.crx'))
    # options.add_extension(os.path.join(os.getcwd(), 'extensions', 'smart-proxy.crx'))
    # options.add_extension(os.path.join(os.getcwd(), 'extensions', 'adguard-adblocker.crx'))
    # options.add_argument('load-extension=' + os.path.join(os.getcwd(), 'extensions', 'XBlocker 1.0.4') + ',' + os.path.join(os.getcwd(), 'extensions', 'XBlocker 1.0.4 - langpack'))
    options.add_experimental_option("detach", True)

    driver = webdriver.Chrome(options=options)
    driver.maximize_window()

    action = webdriver.ActionChains(driver)
    action = webdriver.common.action_chains.ActionChains(driver)

    driver.get("https://www.upwork.com/ab/account-security/login")

    wait = WebDriverWait(driver, 10)

    # Wait for email input field to appear and enter email
    login_username = wait.until(
        EC.visibility_of_element_located((By.ID, "login_username")))
    login_username.send_keys(email)

    # Click on continue button
    login_password_continue = wait.until(
        EC.element_to_be_clickable((By.ID, "login_password_continue")))
    login_password_continue.click()

    # Wait for password input field to appear and enter password
    login_password = wait.until(
        EC.visibility_of_element_located((By.ID, "login_password")))
    login_password.send_keys(pwd)

    # Click on continue button
    login_control_continue = wait.until(
        EC.element_to_be_clickable((By.ID, "login_control_continue")))
    login_control_continue.click()

    # Are your Expert, GET_EXPERIENCE, ...
    sleep(10)
    waitInfinite(lambda: driver.execute_script(
        'document.querySelector("button.air3-btn.air3-btn-primary").click()'))
    sleep(2)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "input[value=\"FREELANCED_BEFORE\"]").click())
    sleep(1)
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]").click())
    sleep(2)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "input[value=\"GET_EXPERIENCE\"]").click())
    sleep(1)
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]").click())
    sleep(2)

    waitInfinite(lambda: driver.execute_script(
        "x=document.querySelectorAll('input[class=\"air3-btn-box-input\"]').length - 2;document.querySelectorAll('input[class=\"air3-btn-box-input\"]')[x].click()"))
    sleep(2)
    waitInfinite(lambda: driver.execute_script(
        "x = document.querySelector('span[data-test=\"checkbox-input\"]'); console.log(x); x.click()"))
    # waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR, "span[data-test=\"checkbox-input\"]").click())
    sleep(1)
    waitInfinite(lambda: driver.execute_script(
        "document.querySelector('button[data-test=\"next-button\"]').click()"))
    sleep(2)
    # waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR, "button[data-test=\"next-button\"]").click())

    # Fill out manually
    waitInfinite(lambda: driver.execute_script(
        'x=document.querySelectorAll("button.mb-3.air3-btn.air3-btn-secondary").length;document.querySelectorAll("button.mb-3.air3-btn.air3-btn-secondary")[x-1].click()'))
    sleep(2)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "input[placeholder=\"Software Engineer | Javascript | iOS\"]").send_keys(profile['professional']))
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]").click())
    sleep(2)
    for i in profile['workXP']:
        addExperience(driver, i)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]").click())

    expFlag = True
    sleep(2)

    for i in profile['education']:
        addEducation(driver, i)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]").click())

    sleep(6)
    # sleep(1)
    try:
        driver.find_element(
            By.CSS_SELECTOR, "button[data-test=\"skip-button\"]").click()
    except:
        pass

    expFlag = True
    sleep(1)
    waitInfinite(lambda: selectDateDropDown(
        "dropdown-label-english", "span.air3-menu-item-text", 2))

    count = 0
    for i in profile['languages']:
        count = addLanguage(driver, i['language'], i['level'], count)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]").click())

    sleep(1)
    inp_skills = driver.find_element(
        By.CSS_SELECTOR, 'input[aria-labelledby="skills-input"]')

    for i in profile['skills']:
        addSkill(driver, inp_skills, i)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]").click())

    sleep(1)
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'textarea[aria-labelledby="overview-label"]').send_keys(profile['overview']))
    sleep(1)
    clickByMouse(driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]"))
    # waitInfinite(lambda: driver.execute_script('document.querySelector("button.air3-btn.air3-btn-primary.mb-0").click()'))

    sleep(3)
    addService(driver, profile['services'])
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]").click())

    sleep(2)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[data-test="currency-input"]').clear())
    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 'input[data-test="currency-input"]').send_keys(str(profile['hourRate'])))

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]").click())

    sleep(2)
    configLast(
        driver,
        profile['location'],
        profile['street'],
        profile['city'],
        profile['birthday'],
        profile['phone'],
        random.choice(profile['avatar'])
    )

    sleep(2)

    waitInfinite(lambda: driver.find_element(By.CSS_SELECTOR,
                 "button[data-test=\"next-button\"]").click())
    # waitInfinite(lambda: driver.execute_script('document.querySelector("button.air3-btn.width-md.m-0.air3-btn-primary").click()'))

    sht.cell(row=p, column=1).value = email
    file_name = email.split("@")[0] + ".xlsx"

    driver.quit()


finish_excel.save(filename=file_name)

print("successfully executed")
